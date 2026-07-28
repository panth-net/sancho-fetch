from __future__ import annotations

import json
import zipfile
from pathlib import Path

import yaml

from sancho.project_export import (
    PROJECT_FOLDER,
    export_records_to_public_outputs,
)


def _make_record(
    base: Path,
    name: str,
    payload,
    provenance: dict,
    original: bytes | None = None,
) -> Path:
    d = base / name
    d.mkdir(parents=True, exist_ok=True)
    (d / "data.json").write_text(json.dumps(payload), encoding="utf-8")
    (d / "provenance.yml").write_text(yaml.safe_dump(provenance), encoding="utf-8")
    if original is not None:
        (d / provenance["original_file"]).write_bytes(original)
    return d


def _export(record_dirs, project: Path, **kw):
    project.mkdir(parents=True, exist_ok=True)
    return export_records_to_public_outputs(
        record_dirs=record_dirs,
        project_root=project,
        workspace_root=project / "ws",
        **kw,
    )


def test_project_folder_name_is_filesystem_safe() -> None:
    assert PROJECT_FOLDER == "sancho-downloads"
    assert " " not in PROJECT_FOLDER


def test_single_tabular_record_exports_flat_xlsx(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "flat",
        {"rows": [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]},
        {"source_kind": "api", "family": "alabama_pop", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".xlsx"
    assert res.primary_path.is_file()
    assert len(res.output_paths) == 1


def test_single_json_record_exports_flat_json(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "blob",
        {"meta": {"a": 1}, "nested": {"deep": [1, 2]}},
        {"source_kind": "api", "family": "blob", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".json"


def _sheet_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    sheet = load_workbook(path).active
    headers = [c.value for c in sheet[1]]
    return [dict(zip(headers, (c.value for c in row))) for row in sheet.iter_rows(min_row=2)]


def test_scalar_subdict_flattens_to_single_table(tmp_path: Path) -> None:
    # One level of all-scalar sub-dicts is lifted into dotted columns
    # (lossless), so no JSON companion is needed.
    rec = _make_record(
        tmp_path / "cache", "nested",
        {"rows": [{"a": 1, "addr": {"city": "NY"}}]},
        {"source_kind": "api", "family": "nested", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".xlsx"
    rows = _sheet_rows(res.primary_path)
    assert rows[0]["addr.city"] == "NY"


def test_xlsx_preserves_code_strings_and_number_types(tmp_path: Path) -> None:
    # The reason tables default to .xlsx: Excel strips leading zeros from
    # CSV code columns (FIPS "01003" -> 1003). String cells must stay text.
    rec = _make_record(
        tmp_path / "cache", "codes",
        {"rows": [{"fips": "01003", "rate": 12.4, "count": 7}]},
        {"source_kind": "api", "family": "codes", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    row = _sheet_rows(res.primary_path)[0]
    assert row["fips"] == "01003"
    assert isinstance(row["rate"], float)
    assert isinstance(row["count"], int)


def test_deep_nested_list_exports_table_and_json(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "nested",
        {"rows": [{"a": 1, "tags": ["x", "y"]}]},
        {"source_kind": "api", "family": "nested", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_dataset_folder"
    exts = sorted(p.suffix for p in res.output_paths)
    assert exts == [".json", ".xlsx"]
    # The nested value is preserved as JSON text in the table cell.
    table_path = next(p for p in res.output_paths if p.suffix == ".xlsx")
    row = _sheet_rows(table_path)[0]
    assert row["tags"] == '["x","y"]'


def test_excel_source_keeps_xlsx_only(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "xl",
        {"rows": [{"a": 1}]},
        {"source_kind": "file", "original_file": "original.xlsx", "family": "cpi", "module_id": "m"},
        original=b"PK\x03\x04excel-bytes",
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".xlsx"
    assert res.primary_path.read_bytes() == b"PK\x03\x04excel-bytes"


def test_parquet_source_exports_csv_and_original(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "pq",
        {"rows": [{"a": 1, "b": 2}]},
        {"source_kind": "file", "original_file": "original.parquet", "family": "overture", "module_id": "m"},
        original=b"PAR1data",
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_dataset_folder"
    exts = sorted(p.suffix for p in res.output_paths)
    assert exts == [".parquet", ".xlsx"]


def test_geojson_kept_native(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "geo",
        {"type": "FeatureCollection", "features": []},
        {"source_kind": "file", "original_file": "original.geojson", "family": "ne", "module_id": "m"},
        original=b'{"type":"FeatureCollection","features":[]}',
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".geojson"


def test_zip_of_files_is_unzipped(tmp_path: Path) -> None:
    buf = tmp_path / "raw.zip"
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("a.csv", "x,y\n1,2\n")
        zf.writestr("b.csv", "p,q\n3,4\n")
    rec = _make_record(
        tmp_path / "cache", "z",
        {"note": "archive"},
        {"source_kind": "file", "original_file": "original.zip", "family": "bundle", "module_id": "m"},
        original=buf.read_bytes(),
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_dataset_folder"
    names = sorted(p.name for p in res.output_paths)
    assert names == ["a.csv", "b.csv"]


def test_shapefile_zip_kept_grouped(tmp_path: Path) -> None:
    buf = tmp_path / "shp.zip"
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("region.shp", "shp")
        zf.writestr("region.dbf", "dbf")
        zf.writestr("region.shx", "shx")
    rec = _make_record(
        tmp_path / "cache", "shp",
        {"note": "shapefile"},
        {"source_kind": "file", "original_file": "original.zip", "family": "boundaries", "module_id": "m"},
        original=buf.read_bytes(),
    )
    res = _export([rec], tmp_path / "proj")
    # Cohesive bundle: kept as the original .zip, not unpacked.
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".zip"


def test_multiple_records_export_to_timestamp_folder(tmp_path: Path) -> None:
    r1 = _make_record(
        tmp_path / "cache", "al",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "alabama", "module_id": "m"},
    )
    r2 = _make_record(
        tmp_path / "cache", "ma",
        {"rows": [{"a": 2}]},
        {"source_kind": "api", "family": "massachusetts", "module_id": "m"},
    )
    res = _export(
        [r1, r2], tmp_path / "proj",
        labels=["alabama-population", "massachusetts-population"],
        unit_sources=["reused_cache", "fetched_api"],
    )
    assert res.mode == "multi_dataset_folder"
    assert res.primary_path.is_dir()
    assert res.reused_count == 1
    assert res.fetched_count == 1
    names = sorted(p.name for p in res.output_paths)
    assert names == ["alabama-population.xlsx", "massachusetts-population.xlsx"]


def test_output_filename_uses_timestamp_and_short_label(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "ts",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "alabama_pop", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj", labels=["alabama-population"])
    # YYYY-MM-DD_HHMMSS_label.csv
    name = res.primary_path.name
    assert name.endswith("_alabama-population.xlsx")
    stamp = name.split("_alabama")[0]
    date_part, _, time_part = stamp.partition("_")
    assert len(date_part) == 10 and date_part[4] == "-" and date_part[7] == "-"
    assert len(time_part) == 6 and time_part.isdigit()


def test_long_label_is_truncated_with_hash(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "long",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "x", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj", labels=["a" * 200])
    # Name shape: YYYY-MM-DD_HHMMSS_<label>.csv
    label_part = res.primary_path.name.split("_", 2)[-1].replace(".xlsx", "")
    assert len(label_part) <= 48
    assert "__" in label_part


def test_duplicate_labels_do_not_overwrite(tmp_path: Path) -> None:
    r1 = _make_record(
        tmp_path / "cache", "d1",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "dup", "module_id": "m"},
    )
    r2 = _make_record(
        tmp_path / "cache", "d2",
        {"rows": [{"a": 2}]},
        {"source_kind": "api", "family": "dup", "module_id": "m"},
    )
    res = _export([r1, r2], tmp_path / "proj", labels=["same", "same"])
    names = sorted(p.name for p in res.output_paths)
    assert names == ["same.xlsx", "same__2.xlsx"]
    assert len(res.output_paths) == 2


def test_csv_is_utf8_sig_and_escapes_delimiters(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "intl",
        {"rows": [{"name": "Müller, GmbH", "note": 'has "quotes"\nand newline'}]},
        {"source_kind": "api", "family": "intl", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj", config={"exports": {"tabular_format": "csv"}})
    raw = res.primary_path.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM
    # Round-trips through the csv reader without corruption.
    text = raw.decode("utf-8-sig")
    import csv as _csv
    import io as _io
    rows = list(_csv.DictReader(_io.StringIO(text)))
    assert rows[0]["name"] == "Müller, GmbH"
    assert rows[0]["note"] == 'has "quotes"\nand newline'


def test_public_export_never_writes_inside_sancho_workspace(tmp_path: Path) -> None:
    rec = _make_record(
        tmp_path / "cache", "x",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "x", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    for path in res.output_paths:
        assert "sancho-workspace" not in str(path)


def test_same_second_same_label_does_not_overwrite(tmp_path: Path) -> None:
    # Two exports in the same second with the same label must not clobber.
    rec = _make_record(
        tmp_path / "cache", "dup",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "pop", "module_id": "m"},
    )
    proj = tmp_path / "proj"
    r1 = _export([rec], proj, labels=["alabama"])
    r2 = _export([rec], proj, labels=["alabama"])
    assert r1.primary_path != r2.primary_path
    assert r1.primary_path.exists() and r2.primary_path.exists()
    files = sorted(p.name for p in (proj / "sancho-downloads").glob("*.xlsx"))
    assert len(files) == 2


def test_api_geojson_is_kept_as_geojson(tmp_path: Path) -> None:
    fc = {"type": "FeatureCollection", "features": [
        {"type": "Feature", "geometry": {"type": "Point", "coordinates": [1, 2]}, "properties": {"n": "A"}}]}
    rec = _make_record(
        tmp_path / "cache", "apigeo", fc,
        {"source_kind": "api", "family": "boundaries", "module_id": "m"},
    )
    res = _export([rec], tmp_path / "proj")
    assert res.mode == "single_file"
    assert res.primary_path.suffix == ".geojson"
    # Round-trips as the same FeatureCollection (geometry not flattened away).
    assert json.loads(res.primary_path.read_text(encoding="utf-8"))["type"] == "FeatureCollection"


def test_missing_record_in_batch_is_skipped(tmp_path: Path) -> None:
    good = _make_record(
        tmp_path / "cache", "good",
        {"rows": [{"a": 1}]},
        {"source_kind": "api", "family": "ok", "module_id": "m"},
    )
    missing = tmp_path / "cache" / "gone"
    res = _export([good, missing], tmp_path / "proj", labels=["ok", "gone"],
                  unit_sources=["fetched_api", "fetched_api"])
    names = [p.name for p in res.output_paths]
    assert any("ok" in n for n in names)
    assert len(res.canonical_record_dirs) == 1


def test_empty_zip_falls_back_to_original(tmp_path: Path) -> None:
    buf = tmp_path / "empty.zip"
    with zipfile.ZipFile(buf, "w"):
        pass
    rec = _make_record(
        tmp_path / "cache", "ez",
        {"x": 1},
        {"source_kind": "file", "original_file": "original.zip", "family": "bundle", "module_id": "m"},
        original=buf.read_bytes(),
    )
    res = _export([rec], tmp_path / "proj")
    names = [p.name for p in res.output_paths]
    assert names and names[0].endswith(".zip")  # kept the original, no empty folder


def test_large_file_is_copied_and_flagged(tmp_path: Path) -> None:
    big = b"x" * 2048
    rec = _make_record(
        tmp_path / "cache", "big",
        {"note": "big"},
        {"source_kind": "file", "original_file": "original.pdf", "family": "report", "module_id": "m"},
        original=big,
    )
    res = export_records_to_public_outputs(
        record_dirs=[rec],
        project_root=tmp_path / "proj",
        workspace_root=tmp_path / "ws",
        config={"exports": {"large_file_warn_bytes": 1024}},
    )
    assert res.primary_path.exists()
    assert res.large_files
    assert res.large_files[0]["bytes"] == 2048
