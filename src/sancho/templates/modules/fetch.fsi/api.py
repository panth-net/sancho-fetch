from __future__ import annotations

import io
import re
import tempfile
from pathlib import Path
from typing import Any

import requests

# ---------------------------------------------------------------------------
# FSI publishes XLSX reports at this index page.
# ---------------------------------------------------------------------------

INDEX_URL = "https://fragilestatesindex.org/excel/"


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------


def build_source_url(*, year: int | None) -> str:
    """Return the index page URL (actual download link is discovered at runtime)."""
    return INDEX_URL


_BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)


def _discover_xlsx_urls(timeout: float) -> list[tuple[int, str]]:
    """Scrape the FSI Excel index page for .xlsx download links.

    Returns a list of (year, url) tuples sorted newest-first.
    """
    resp = requests.get(INDEX_URL, headers={"User-Agent": _BROWSER_UA}, timeout=timeout)
    resp.raise_for_status()
    links = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', resp.text, re.IGNORECASE)
    results: list[tuple[int, str]] = []
    for link in links:
        year_match = re.search(r'20\d{2}', link)
        if year_match:
            results.append((int(year_match.group()), link))
    results.sort(key=lambda x: x[0], reverse=True)
    return results


def _pick_xlsx_url(
    *, year: int | None, timeout: float,
) -> tuple[int, str]:
    """Select the XLSX URL for the requested year, or the latest available."""
    available = _discover_xlsx_urls(timeout=timeout)
    if not available:
        raise RuntimeError(
            f"No .xlsx download links found on {INDEX_URL}. "
            "The page structure may have changed."
        )

    if year is not None:
        for avail_year, url in available:
            if avail_year == year:
                return (avail_year, url)
        years_found = [str(y) for y, _ in available]
        raise RuntimeError(
            f"FSI XLSX for year {year} not found. "
            f"Available years: {', '.join(years_found)}"
        )

    # Latest
    return available[0]


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

_SKIP_COLS = {"country", "year", "rank", "total"}


def _normalise(text: Any) -> str:
    return str(text).strip().lower().replace("_", " ") if text else ""


def _parse_xlsx(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    """Parse the FSI XLSX into long-format rows.

    Each row in the spreadsheet has Country, Year, Rank, Total, and individual
    metric columns.  We unpivot the metric columns into long format:
    {"country": ..., "year": ..., "metric": ..., "value": ...}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        raise RuntimeError(f"Empty sheet in {source_url}")

    # Build column index
    col_names = [_normalise(h) for h in header]
    country_idx: int | None = None
    year_idx: int | None = None
    metric_indices: list[tuple[int, str]] = []

    for idx, name in enumerate(col_names):
        if name in ("country", "country / territory", "country/territory"):
            country_idx = idx
        elif name == "year":
            year_idx = idx
        elif name and name not in _SKIP_COLS:
            metric_indices.append((idx, str(header[idx]).strip()))

    if country_idx is None:
        wb.close()
        raise RuntimeError(
            f"'Country' column not found in header: {list(header)}"
        )

    parsed_rows: list[dict[str, Any]] = []
    for data_row in rows_iter:
        country_val = data_row[country_idx]
        if country_val is None:
            continue

        country_str = str(country_val).strip()

        # Apply country filter (case-insensitive substring match)
        if country and country not in country_str.upper():
            continue

        row_year = (
            int(data_row[year_idx])
            if year_idx is not None and data_row[year_idx] is not None
            else resolved_year
        )

        for m_idx, m_name in metric_indices:
            raw_val = data_row[m_idx] if m_idx < len(data_row) else None
            if raw_val is None:
                continue
            try:
                value = float(raw_val)
            except (ValueError, TypeError):
                continue
            parsed_rows.append({
                "country": country_str,
                "year": row_year,
                "metric": m_name,
                "value": value,
            })

    wb.close()
    return {
        "source_url": source_url,
        "rows": parsed_rows,
        "row_count": len(parsed_rows),
    }


# ---------------------------------------------------------------------------
# Fallback when openpyxl is not installed
# ---------------------------------------------------------------------------


def _fallback_without_openpyxl(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
) -> dict[str, Any]:
    tmp_dir = Path(tempfile.gettempdir()) / "sancho_fsi"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out_path = tmp_dir / f"FSI{resolved_year}.xlsx"
    out_path.write_bytes(xlsx_bytes)

    return {
        "source_url": source_url,
        "rows": [],
        "row_count": 0,
        "saved_xlsx_path": str(out_path),
        "note": (
            "openpyxl is not installed. The raw XLSX has been saved to "
            f"'{out_path}'. Install openpyxl (`pip install openpyxl`) and "
            "re-run to get parsed rows."
        ),
    }


# ---------------------------------------------------------------------------
# Public fetch entry point
# ---------------------------------------------------------------------------


def fetch_fsi(
    runtime_http: dict[str, Any],
    year: int | None = None,
    country: str | None = None,
) -> dict[str, Any]:
    """Download the FSI XLSX and parse it into long-format row dicts."""
    timeout = runtime_http.get("timeout_seconds", 30)
    resolved_year, xlsx_url = _pick_xlsx_url(year=year, timeout=timeout)

    resp = requests.get(xlsx_url, headers={"User-Agent": _BROWSER_UA}, timeout=timeout)
    resp.raise_for_status()
    xlsx_bytes = resp.content

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return _fallback_without_openpyxl(
            xlsx_bytes=xlsx_bytes,
            source_url=xlsx_url,
            resolved_year=resolved_year,
        )

    return _parse_xlsx(
        xlsx_bytes=xlsx_bytes,
        source_url=xlsx_url,
        resolved_year=resolved_year,
        country=country,
    )
