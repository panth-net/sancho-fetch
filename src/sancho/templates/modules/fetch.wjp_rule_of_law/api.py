from __future__ import annotations

import io
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests

# ---------------------------------------------------------------------------
# URL pattern -- WJP publishes XLSX historical data files each year.
# ---------------------------------------------------------------------------

_BASE_URL = (
    "https://worldjusticeproject.org/rule-of-law-index/downloads/"
    "{year}_wjp_rule_of_law_index_HISTORICAL_DATA_FILE.xlsx"
)
_PROBE_YEARS_BACK = 12

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _url_for_year(year: int) -> str:
    return _BASE_URL.format(year=year)


def _discover_latest_year(timeout: int = 15) -> int:
    """HEAD-probe backwards from the current year to find the newest XLSX report.

    WJP's CDN returns 200 + text/html for missing years (SPA fallback), so we
    must check Content-Type -- not just the status code.
    """
    current_year = datetime.now(timezone.utc).year
    for y in range(current_year, current_year - _PROBE_YEARS_BACK, -1):
        url = _url_for_year(y)
        try:
            resp = requests.head(url, timeout=timeout, allow_redirects=True)
            if resp.status_code >= 400:
                continue
            ct = resp.headers.get("Content-Type", "").lower()
            # Only accept XLSX/binary responses -- reject text/html SPA fallbacks
            if "spreadsheet" in ct or "octet-stream" in ct or "excel" in ct:
                return y
        except requests.RequestException:
            continue
    raise RuntimeError(
        f"Could not locate a WJP Rule of Law XLSX report for any year in "
        f"[{current_year - _PROBE_YEARS_BACK + 1}..{current_year}]."
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_source_url(*, year: int | None = None) -> str:
    """Return the download URL, discovering the latest year if needed."""
    resolved = year if year is not None else _discover_latest_year()
    return _url_for_year(resolved)


def fetch_wjp_rule_of_law(
    runtime_http: dict[str, Any],
    year: int | None = None,
    country: str | None = None,
) -> dict[str, Any]:
    """Download the WJP Rule of Law XLSX and parse it into row dicts.

    If *openpyxl* is available the XLSX is parsed in-memory.  Otherwise the
    raw bytes are saved to a temp file and metadata is returned with
    instructions for the user.
    """
    timeout = runtime_http.get("timeout_seconds", 30)
    resolved_year = year if year is not None else _discover_latest_year(timeout=timeout)
    source_url = _url_for_year(resolved_year)

    from sancho.runtime.net import download_file as _dl
    result = _dl(
        source_url,
        dest_dir=Path(tempfile.gettempdir()) / "sancho_wjp",
        filename=f"WJP_{resolved_year}.xlsx",
        expected_magic=b"PK",
        timeout=float(timeout),
    )
    xlsx_bytes = result.path.read_bytes()

    try:
        import openpyxl  # noqa: F811
    except ImportError:
        return _fallback_without_openpyxl(
            xlsx_bytes=xlsx_bytes,
            source_url=source_url,
            resolved_year=resolved_year,
            country=country,
        )

    return _parse_xlsx(
        xlsx_bytes=xlsx_bytes,
        source_url=source_url,
        resolved_year=resolved_year,
        country=country,
    )


# ---------------------------------------------------------------------------
# XLSX parsing (openpyxl available)
# ---------------------------------------------------------------------------

_SHEET_NAMES = ["Historical Data", "WJP Historical", "Data", "Sheet1"]

_COL_ALIASES: dict[str, list[str]] = {
    "country": ["country", "country/territory", "country / territory"],
    "country_iso3": [
        "country code (iso3)",
        "country code",
        "iso3",
        "iso",
        "code",
    ],
    "year": ["year"],
}

# Columns that are NOT metric scores (used to identify metric columns).
_NON_METRIC_FIELDS = {"country", "country_iso3", "year"}


def _normalise(text: Any) -> str:
    return str(text).strip().lower().replace("_", " ") if text else ""


def _resolve_columns(header_cells: tuple) -> tuple[dict[str, int], list[tuple[int, str]]]:
    """Map logical field names to 0-based column indices.

    Returns (known_map, metric_columns) where metric_columns is a list of
    (index, original_header_text) for every column not matched to a known
    field.
    """
    known: dict[str, int] = {}
    matched_indices: set[int] = set()
    for idx, cell in enumerate(header_cells):
        norm = _normalise(cell)
        for field, aliases in _COL_ALIASES.items():
            if norm in aliases and field not in known:
                known[field] = idx
                matched_indices.add(idx)

    metric_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_cells):
        if idx not in matched_indices and cell is not None:
            header_text = str(cell).strip()
            if header_text:
                metric_cols.append((idx, header_text))

    return known, metric_cols


def _parse_xlsx(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True,
    )

    # Dynamic sheet selection: try known names, then fall back.
    ws = None
    for name in _SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)

    # First non-empty row is treated as the header.
    header: tuple | None = None
    for row in rows_iter:
        if any(cell is not None for cell in row):
            header = row
            break

    if header is None:
        wb.close()
        raise RuntimeError(
            f"No header row found in sheet '{ws.title}' of {source_url}"
        )

    col_map, metric_cols = _resolve_columns(header)
    required = {"country"}
    missing = required - col_map.keys()
    if missing:
        wb.close()
        raise RuntimeError(
            f"Required columns {missing} not found in header: {header}"
        )

    # Unpivot: each data row x each metric column = one output row.
    parsed_rows: list[dict[str, Any]] = []
    for data_row in rows_iter:
        country_val = data_row[col_map["country"]] if "country" in col_map else None
        iso3_val = (
            data_row[col_map["country_iso3"]]
            if "country_iso3" in col_map
            else None
        )
        year_val = (
            data_row[col_map["year"]] if "year" in col_map else None
        )

        if country_val is None and iso3_val is None:
            continue  # skip blank rows

        # Apply country filter (match on ISO3 or country name).
        if country:
            iso3_upper = str(iso3_val).upper() if iso3_val else ""
            country_upper = str(country_val).upper() if country_val else ""
            if country != iso3_upper and country != country_upper:
                continue

        # Year column can be int ("2024"), float (2024.0), range string ("2012-2013").
        # For ranges, take the end year; for numbers, cast to int.
        if year_val is None:
            row_year = resolved_year
        elif isinstance(year_val, (int, float)):
            row_year = int(year_val)
        else:
            year_str = str(year_val).strip()
            if "-" in year_str:
                parts = year_str.split("-")
                try:
                    row_year = int(parts[-1].strip())
                except ValueError:
                    row_year = resolved_year
            else:
                try:
                    row_year = int(float(year_str))
                except ValueError:
                    row_year = resolved_year
        row_country = str(country_val).strip() if country_val else None
        row_iso3 = str(iso3_val).strip() if iso3_val else None

        for col_idx, metric_name in metric_cols:
            value = data_row[col_idx]
            if value is None:
                continue
            try:
                float_val = float(value)
            except (TypeError, ValueError):
                continue
            parsed_rows.append({
                "country": row_country,
                "country_iso3": row_iso3,
                "year": row_year,
                "metric": metric_name,
                "value": float_val,
            })

    wb.close()

    return {
        "source_url": source_url,
        "rows": parsed_rows,
        "row_count": len(parsed_rows),
    }


# ---------------------------------------------------------------------------
# Fallback when openpyxl is not installed
# ---------------------------------------------------------------------------


def _fallback_without_openpyxl(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    tmp_dir = Path(tempfile.gettempdir()) / "sancho_wjp_rule_of_law"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out_path = tmp_dir / f"WJP_{resolved_year}_historical_data.xlsx"
    out_path.write_bytes(xlsx_bytes)

    return {
        "source_url": source_url,
        "rows": [],
        "row_count": 0,
        "saved_xlsx_path": str(out_path),
        "note": (
            "openpyxl is not installed. The raw XLSX has been saved to "
            f"'{out_path}'. Install openpyxl (`pip install openpyxl`) and "
            "re-run to get parsed rows."
        ),
    }
