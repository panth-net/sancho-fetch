from __future__ import annotations

import io
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests

_BASE_URL = "https://images.transparencycdn.org/images/CPI{year}-Results-and-trends.xlsx"
_PROBE_YEARS_BACK = 12


def _url_for_year(year: int) -> str:
    return _BASE_URL.format(year=year)


def _discover_latest_year(timeout: int = 15) -> int:
    """HEAD-probe backwards from the current year to find the newest report."""
    current_year = datetime.now(timezone.utc).year
    for y in range(current_year, current_year - _PROBE_YEARS_BACK, -1):
        url = _url_for_year(y)
        try:
            resp = requests.head(url, timeout=timeout, allow_redirects=True)
            if resp.status_code < 400:
                return y
        except requests.RequestException:
            continue
    raise RuntimeError(
        f"Could not locate a CPI XLSX report for any year in "
        f"[{current_year - _PROBE_YEARS_BACK + 1}..{current_year}]."
    )


def build_source_url(*, year: int | None = None) -> str:
    """Return the download URL, discovering the latest year if needed."""
    resolved = year if year is not None else _discover_latest_year()
    return _url_for_year(resolved)


def fetch_cpi(
    runtime_http: dict[str, Any],
    year: int | None = None,
    country: str | None = None,
) -> dict[str, Any]:
    """Download the CPI XLSX and parse it into row dicts.

    If *openpyxl* is available the XLSX is parsed in-memory.  Otherwise the
    raw bytes are saved to a temp file and metadata is returned with
    instructions for the user.
    """
    from sancho.runtime.net import download_file as _dl
    timeout = runtime_http.get("timeout_seconds", 30)
    resolved_year = year if year is not None else _discover_latest_year(timeout=timeout)
    source_url = _url_for_year(resolved_year)

    result = _dl(
        source_url,
        dest_dir=Path(tempfile.gettempdir()) / "sancho_ti_cpi",
        filename=f"CPI{resolved_year}.xlsx",
        expected_magic=b"PK",
        timeout=float(timeout),
    )
    xlsx_bytes = result.path.read_bytes()

    try:
        import openpyxl  # noqa: F811
    except ImportError:
        return _fallback_without_openpyxl(
            xlsx_bytes=xlsx_bytes,
            source_url=source_url,
            resolved_year=resolved_year,
            country=country,
        )

    return _parse_xlsx(
        xlsx_bytes=xlsx_bytes,
        source_url=source_url,
        resolved_year=resolved_year,
        country=country,
    )


_SHEET_NAMES = ["CPI Historical", "CPI Timeseries", "Results", "CPI Results"]
_COL_ALIASES: dict[str, list[str]] = {
    "iso3": ["iso3", "iso", "country_iso3"],
    "country": ["country", "country / territory", "country/territory"],
    "year": ["year"],
    "cpi_score": ["cpi score", "cpiscore", "cpi_score", "score"],
}


def _normalise(text: Any) -> str:
    return str(text).strip().lower().replace("_", " ") if text else ""


def _resolve_columns(header_cells: tuple) -> dict[str, int]:
    """Map logical field names to 0-based column indices."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        norm = _normalise(cell)
        for field, aliases in _COL_ALIASES.items():
            if norm in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _rows_to_output(
    *,
    raw_rows: list[list[Any]],
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    """Turn a raw list-of-lists (from strict-OOXML parser) into the output shape."""
    header: list | None = None
    for row in raw_rows:
        norms = [_normalise(c) for c in row if c is not None]
        if any(n in ("iso3", "iso", "country", "country / territory") for n in norms):
            header = row
            break
    if header is None:
        raise RuntimeError(f"No header row found in strict-OOXML sheet at {source_url}")

    col_map = _resolve_columns(tuple(header))
    data_rows = raw_rows[raw_rows.index(header) + 1:]
    parsed: list[dict[str, Any]] = []
    for data_row in data_rows:
        if not data_row:
            continue
        iso3_val = data_row[col_map["iso3"]] if "iso3" in col_map and col_map["iso3"] < len(data_row) else None
        country_val = data_row[col_map["country"]] if "country" in col_map and col_map["country"] < len(data_row) else None
        year_val = data_row[col_map["year"]] if "year" in col_map and col_map["year"] < len(data_row) else None
        score_val = data_row[col_map["cpi_score"]] if "cpi_score" in col_map and col_map["cpi_score"] < len(data_row) else None
        if iso3_val is None and country_val is None:
            continue
        if country and str(iso3_val).upper() != country and str(country_val).upper() != country:
            continue
        try:
            score = float(score_val) if score_val is not None else None
        except (ValueError, TypeError):
            score = None
        parsed.append({
            "iso3": str(iso3_val).strip() if iso3_val else None,
            "country": str(country_val).strip() if country_val else None,
            "year": int(year_val) if isinstance(year_val, (int, float)) else resolved_year,
            "cpi_score": score,
        })
    return {"source_url": source_url, "rows": parsed, "row_count": len(parsed)}


def _parse_strict_ooxml_rows(xlsx_bytes: bytes) -> list[list[Any]]:
    """Fallback XML parser for strict-OOXML XLSX files (openpyxl doesn't read sheets).

    TI CPI 2024 uses the strict OOXML namespace which openpyxl doesn't handle,
    resulting in empty sheetnames. We parse sheet1 directly from the zip.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        # Build shared-strings lookup
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_xml = zf.read("xl/sharedStrings.xml")
            ss_root = ET.fromstring(ss_xml)
            for si in ss_root:
                text_parts = [t.text or "" for t in si.iter() if t.tag.endswith("}t")]
                shared_strings.append("".join(text_parts))

        # Try each sheet, looking for one with CPI data
        for idx in range(1, 10):
            sheet_name = f"xl/worksheets/sheet{idx}.xml"
            if sheet_name not in zf.namelist():
                break
            sheet_xml = zf.read(sheet_name)
            root = ET.fromstring(sheet_xml)
            rows_out: list[list[Any]] = []
            for row_el in root.iter():
                if not row_el.tag.endswith("}row"):
                    continue
                row_cells: list[Any] = []
                for cell in row_el:
                    if not cell.tag.endswith("}c"):
                        continue
                    t = cell.get("t", "n")
                    val_el = next((c for c in cell if c.tag.endswith("}v")), None)
                    inline_el = next((c for c in cell if c.tag.endswith("}is")), None)
                    if t == "s" and val_el is not None:
                        try:
                            row_cells.append(shared_strings[int(val_el.text)])
                        except (ValueError, IndexError):
                            row_cells.append(val_el.text)
                    elif t == "inlineStr" and inline_el is not None:
                        row_cells.append("".join(t.text or "" for t in inline_el.iter() if t.tag.endswith("}t")))
                    elif val_el is not None:
                        try:
                            row_cells.append(float(val_el.text) if "." in val_el.text else int(val_el.text))
                        except (ValueError, TypeError):
                            row_cells.append(val_el.text)
                    else:
                        row_cells.append(None)
                rows_out.append(row_cells)
            # If this sheet has CPI-like data, return it
            for r in rows_out[:20]:
                norms = [_normalise(c) for c in r if c is not None]
                if any(n in ("iso3", "iso", "country", "country / territory") for n in norms):
                    return rows_out
            if rows_out and idx == 1:
                return rows_out  # First sheet as fallback
    return []


def _parse_xlsx(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    # Fallback: strict-OOXML files (like TI CPI 2024) appear empty to openpyxl.
    if not wb.sheetnames:
        raw_rows = _parse_strict_ooxml_rows(xlsx_bytes)
        if not raw_rows:
            raise RuntimeError(f"XLSX has no readable sheets at {source_url}")
        return _rows_to_output(
            raw_rows=raw_rows,
            source_url=source_url,
            resolved_year=resolved_year,
            country=country,
        )

    # Dynamic sheet selection: try known names, then find by column content.
    ws = None
    for name in _SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        # Scan sheets for one with expected columns
        for name in wb.sheetnames:
            candidate = wb[name]
            for row in candidate.iter_rows(max_row=10, values_only=True):
                norms = [_normalise(c) for c in row if c is not None]
                if any(n in ("iso3", "iso", "country") for n in norms):
                    ws = candidate
                    break
            if ws is not None:
                break
    if ws is None:
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            raise RuntimeError(f"XLSX has no sheets. Available: {wb.sheetnames}")

    rows_iter = ws.iter_rows(values_only=True)

    # Scan for header row by finding row with expected column names.
    header: tuple | None = None
    for row in rows_iter:
        norms = [_normalise(c) for c in row if c is not None]
        if any(n in ("iso3", "iso", "country", "country / territory") for n in norms):
            header = row
            break

    if header is None:
        wb.close()
        raise RuntimeError(
            f"No header row with expected columns found in sheet "
            f"'{ws.title}' of {source_url}. Sheets: {wb.sheetnames}"
        )

    col_map = _resolve_columns(header)
    required = {"iso3", "country"}
    missing = required - col_map.keys()
    if missing:
        wb.close()
        raise RuntimeError(
            f"Required columns {missing} not found in header: {header}"
        )

    parsed_rows: list[dict[str, Any]] = []
    for data_row in rows_iter:
        iso3_val = data_row[col_map["iso3"]] if "iso3" in col_map else None
        country_val = data_row[col_map["country"]] if "country" in col_map else None
        year_val = data_row[col_map.get("year", -1)] if "year" in col_map else None
        score_val = (
            data_row[col_map["cpi_score"]] if "cpi_score" in col_map else None
        )

        if iso3_val is None and country_val is None:
            continue  # skip blank rows

        # Apply country filter
        if country and str(iso3_val).upper() != country and str(country_val).upper() != country:
            continue

        parsed_rows.append({
            "iso3": str(iso3_val).strip() if iso3_val else None,
            "country": str(country_val).strip() if country_val else None,
            "year": int(year_val) if year_val is not None else resolved_year,
            "cpi_score": float(score_val) if score_val is not None else None,
        })

    wb.close()

    return {
        "source_url": source_url,
        "rows": parsed_rows,
        "row_count": len(parsed_rows),
    }


def _fallback_without_openpyxl(
    *,
    xlsx_bytes: bytes,
    source_url: str,
    resolved_year: int,
    country: str | None,
) -> dict[str, Any]:
    tmp_dir = Path(tempfile.gettempdir()) / "sancho_ti_cpi"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out_path = tmp_dir / f"CPI{resolved_year}_Results.xlsx"
    out_path.write_bytes(xlsx_bytes)

    return {
        "source_url": source_url,
        "rows": [],
        "row_count": 0,
        "saved_xlsx_path": str(out_path),
        "note": (
            "openpyxl is not installed. The raw XLSX has been saved to "
            f"'{out_path}'. Install openpyxl (`pip install openpyxl`) and "
            "re-run to get parsed rows."
        ),
    }
