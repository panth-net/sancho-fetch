from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import requests

URL_TEMPLATE = "https://dashboards.sdgindex.org/static/downloads/files/SDR{year}-data.xlsx"
SHEET_NAME = "Backdated SDG Index"


def build_source_url(*, year: int | None) -> str:
    effective_year = year or _discover_latest_year()
    return URL_TEMPLATE.format(year=effective_year)


def _discover_latest_year() -> int:
    """HEAD-probe back 12 years to find the most recent published dataset."""
    current_year = datetime.now().year
    for candidate in range(current_year, current_year - 12, -1):
        url = URL_TEMPLATE.format(year=candidate)
        try:
            resp = requests.head(url, timeout=10, allow_redirects=True)
            if resp.status_code == 200:
                return candidate
        except requests.RequestException:
            continue
    return 2024


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse the XLSX content using openpyxl (fail-fast if missing)."""
    try:
        import openpyxl  # noqa: WPS433
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse SDG Index XLSX files. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        # Fall back to first sheet if expected name is absent
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        wb.close()
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        record: dict[str, Any] = {}
        for idx, cell_value in enumerate(row):
            if idx < len(headers):
                record[headers[idx]] = cell_value
        records.append(record)

    wb.close()
    return records


def _normalize_row(record: dict[str, Any]) -> dict[str, Any]:
    """Ensure each row has a consistent country_iso3 key."""
    iso3 = ""
    for key in ("id", "Id", "ID", "iso3", "ISO3", "Code", "code"):
        val = record.get(key)
        if val and str(val).strip():
            iso3 = str(val).strip().upper()
            break
    row: dict[str, Any] = {"country_iso3": iso3}
    for key, val in record.items():
        if val is not None:
            row[key] = val
    return row


def fetch_sdg_index(
    *,
    runtime_http: dict[str, Any],
    year: int | None,
    country: str | None,
) -> dict[str, Any]:
    url = build_source_url(year=year)
    timeout = float(runtime_http.get("timeout_seconds", 60))

    resp = requests.get(url, timeout=timeout)
    resp.raise_for_status()

    records = _parse_xlsx(resp.content)
    rows: list[dict[str, Any]] = []
    for record in records:
        normalized = _normalize_row(record)
        if country and normalized.get("country_iso3") != country:
            continue
        rows.append(normalized)

    return {"source_url": url, "rows": rows, "row_count": len(rows)}
